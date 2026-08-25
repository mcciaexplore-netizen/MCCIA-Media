import json, re
from copy import deepcopy
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
CURRENT = ROOT / "app" / "records.json"
INVENTORY = ROOT / "source_inventory.json"
RESEARCH = Path(r"C:\Users\Aarushi Gupta\Downloads\1787634550790-Prashant_Girbane_Comprehensive_Research_2018_2026.xlsx")
INTERNAL = Path(r"C:\Users\Aarushi Gupta\Downloads\Mccia Media.xlsx")

def text(v): return str(v).strip() if v is not None else ""
def norm(v): return re.sub(r"[^a-z0-9]+", " ", text(v).lower()).strip()
def canonical(url):
    if not url: return ""
    try:
        p = urlsplit(text(url))
        path = re.sub(r"/+", "/", p.path).rstrip("/")
        return urlunsplit((p.scheme.lower(), p.netloc.lower().removeprefix("www."), path, "", ""))
    except Exception: return text(url).rstrip("/")
def broad_type(value):
    v=norm(value)
    if "video" in v or "tv interview" in v: return "Video"
    if "pdf" in v or "publication" in v or "report" in v or "brochure" in v: return "PDF"
    if "image" in v or "photo" in v: return "Image"
    if "social" in v or "linkedin" in v or "twitter" in v: return "Social"
    if any(k in v for k in ("article","interview","release","news","op ed")): return "Article"
    return "Other"
def date_value(v):
    if hasattr(v,"strftime"): return v.strftime("%Y-%m-%d")
    s=text(v)
    m=re.match(r"(20\d{2})[-/](\d{1,2})(?:[-/](\d{1,2}))?",s)
    return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3) or 1):02d}" if m else ""
def year_value(v, date=""):
    try: return int(v)
    except Exception:
        m=re.search(r"20\d{2}",date or text(v)); return int(m.group()) if m else 0
def domain(url):
    try: return urlsplit(url).netloc.removeprefix("www.")
    except Exception: return "Web source"

records = json.loads(CURRENT.read_text(encoding="utf-8"))
for r in records:
    r["sourceDataset"]="Existing dashboard"
    r["canonicalUrl"]=canonical(r.get("url"))
    r["duplicateCount"]=1
    r["mergeNotes"]="Baseline structured record"

by_url={r["canonicalUrl"]:r for r in records if r["canonicalUrl"]}
by_key={(norm(r.get("title")),r.get("date","")[:10]):r for r in records}
stats={"baseline":len(records),"research_added":0,"research_matched":0,"pdf_added":0,"pdf_matched":0,"internal_added":0,"internal_matched":0}

wb=openpyxl.load_workbook(RESEARCH,data_only=True,read_only=True)
ws=wb["Prashant Girbane Sources"]
rows=list(ws.iter_rows(values_only=True)); headers=[text(v) for v in rows[0]]
for row in rows[1:]:
    d=dict(zip(headers,row)); url=text(d.get("Source URL")); cu=canonical(url); dt=date_value(d.get("Date Published")); title=text(d.get("Title/Headline"))
    found=by_url.get(cu) if cu else by_key.get((norm(title),dt))
    if found:
        found["duplicateCount"]+=1; found["sourceDataset"] += "; Comprehensive research workbook"; stats["research_matched"]+=1
        if not found.get("description"): found["description"]=text(d.get("Summary/Key Points"))
        continue
    rec={"id":"","date":dt,"year":year_value(d.get("Year"),dt),"type":broad_type(d.get("Content Type")),"format":text(d.get("Content Type")) or "Web source","publisher":text(d.get("Source Name/Publisher")) or domain(url),"title":title or "Untitled source","language":text(d.get("Language")) or "Unknown","presence":"Named / occurrence reported","topic":"General","description":text(d.get("Summary/Key Points")),"status":text(d.get("Verification Status")) or "Partially verified","url":url or None,"mediaUrl":None,"notes":"Imported from comprehensive research workbook; review content against source","sourceDataset":"Comprehensive research workbook","canonicalUrl":cu,"duplicateCount":1,"mergeNotes":"New source not present in dashboard baseline"}
    records.append(rec); stats["research_added"]+=1
    if cu: by_url[cu]=rec
    by_key[(norm(rec["title"]),rec["date"])]=rec

inventory=json.loads(INVENTORY.read_text(encoding="utf-8"))
for url in inventory["pdf"]["urls"]:
    cu=canonical(url)
    if cu in by_url:
        by_url[cu]["duplicateCount"]+=1; by_url[cu]["sourceDataset"] += "; PDF report source index"; stats["pdf_matched"]+=1; continue
    slug=Path(urlsplit(url).path).name or domain(url)
    title=re.sub(r"[-_]+"," ",slug).strip()[:180] or "Source listed in PDF report"
    rec={"id":"","date":"","year":year_value(url),"type":broad_type(url),"format":"Source index URL","publisher":domain(url),"title":title,"language":"Unknown","presence":"MCCIA / DG relevance reported in source index","topic":"General","description":"Direct URL extracted from the comprehensive PDF report source index. Page-level context requires review.","status":"Partially verified","url":url,"mediaUrl":None,"notes":"Imported from PDF source index","sourceDataset":"PDF report source index","canonicalUrl":cu,"duplicateCount":1,"mergeNotes":"URL appeared only in PDF source index"}
    records.append(rec); by_url[cu]=rec; stats["pdf_added"]+=1

wb2=openpyxl.load_workbook(INTERNAL,data_only=True,read_only=True)
news=wb2["News"]; rows=list(news.iter_rows(values_only=True)); headers=[text(v) for v in rows[0]]
for row in rows[1:]:
    d=dict(zip(headers,row)); title=text(d.get("News Headline")); typ=text(d.get("Type")); pub=text(d.get("Publication")); dt=date_value(d.get("Date"))
    if not title or (not typ and not pub and not dt): continue
    key=(norm(title),dt)
    if key in by_key:
        by_key[key]["duplicateCount"]+=1; by_key[key]["sourceDataset"] += "; Internal media tracker"; stats["internal_matched"]+=1; continue
    quote=text(d.get("DG Quote")); published=bool(pub and quote.lower() not in ("not published","not published (to be re shared next week)"))
    rec={"id":"","date":dt,"year":year_value("",dt),"type":broad_type(typ),"format":typ or "Internal media lead","publisher":pub or "Publisher not recorded","title":title,"language":text(d.get("Language")) or "Unknown","presence":"DG quote recorded" if quote and quote!='NA' else "MCCIA-related item","topic":"Internal media tracking","description":quote if quote and quote!='NA' else "Item listed in the MCCIA internal media tracker; no public source URL supplied.","status":"Unverified","url":None,"mediaUrl":None,"notes":"Internal tracker item; public source link missing" if published else "Planned, proposed or publication status unclear","sourceDataset":"Internal media tracker","canonicalUrl":"","duplicateCount":1,"mergeNotes":"Linkless internal record retained as a lead"}
    records.append(rec); by_key[key]=rec; stats["internal_added"]+=1

pubs=wb2["Publications"]
for row in list(pubs.iter_rows(values_only=True))[1:]:
    title=text(row[1] if len(row)>1 else "")
    if not title: continue
    key=(norm(title),"")
    if key in by_key: by_key[key]["duplicateCount"]+=1; stats["internal_matched"]+=1; continue
    rec={"id":"","date":"","year":0,"type":"PDF","format":"Publication inventory","publisher":"MCCIA","title":title,"language":"Unknown","presence":"MCCIA publication; DG occurrence not yet checked","topic":"MCCIA publication","description":"Publication listed in MCCIA's internal dissemination tracker.","status":"Unverified","url":None,"mediaUrl":None,"notes":"Confirm public publication URL and inspect for DG name/photo","sourceDataset":"Internal publications tracker","canonicalUrl":"","duplicateCount":1,"mergeNotes":"Linkless publication lead retained for research"}
    records.append(rec); by_key[key]=rec; stats["internal_added"]+=1

records.sort(key=lambda r:(r.get("date") or "0000",r.get("title") or ""),reverse=True)
for idx,r in enumerate(records,1):
    r["id"]=f"PG{idx:03d}"
    r.pop("canonicalUrl",None)

(ROOT/"app"/"records.json").write_text(json.dumps(records,ensure_ascii=False,indent=2),encoding="utf-8")
comparison={"stats":stats,"final_records":len(records),"with_url":sum(bool(r.get("url")) for r in records),"without_url":sum(not r.get("url") for r in records),"verified":sum(r.get("status")=="Verified" for r in records),"partial":sum(r.get("status")=="Partially verified" for r in records),"unverified":sum(r.get("status")=="Unverified" for r in records),"duplicates_merged":sum(max(0,int(r.get("duplicateCount",1))-1) for r in records)}
(ROOT/"comparison_report.json").write_text(json.dumps(comparison,ensure_ascii=False,indent=2),encoding="utf-8")
print(json.dumps(comparison,ensure_ascii=False,indent=2))
