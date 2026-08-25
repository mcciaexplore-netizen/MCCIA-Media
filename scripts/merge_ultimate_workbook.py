import json, re
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r'C:\Users\Aarushi Gupta\Downloads\1787635770532-Prashant_Girbane_MCCIA_ULTIMATE_Research.xlsx')
TARGET = ROOT / 'app' / 'records.json'
COMPARE = ROOT / 'comparison_report.json'
LABEL = 'Prashant Girbane MCCIA ULTIMATE Research workbook'

def text(value):
    return str(value).strip() if value is not None else ''

def norm(value):
    return re.sub(r'[^a-z0-9]+', ' ', text(value).lower()).strip()

def canonical(url):
    value = text(url)
    if not value:
        return ''
    try:
        parts = urlsplit(value if '://' in value else 'https://' + value)
        host = parts.netloc.lower().removeprefix('www.')
        path = re.sub(r'/+', '/', parts.path).rstrip('/')
        return urlunsplit(('', host, path, '', ''))
    except ValueError:
        return value.lower().split('?', 1)[0].split('#', 1)[0].rstrip('/')

def date_value(value):
    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d')
    value = text(value)
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%B %d, %Y', '%b %d, %Y'):
        try:
            return datetime.strptime(value, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    match = re.search(r'\b(20\d{2})[-/](\d{1,2})[-/](\d{1,2})\b', value)
    return f'{match.group(1)}-{int(match.group(2)):02d}-{int(match.group(3)):02d}' if match else ''

def year_value(year, parsed_date):
    match = re.search(r'\b(20\d{2})\b', text(year))
    return int(match.group(1)) if match else (int(parsed_date[:4]) if parsed_date else None)

def broad_type(source_type):
    value = norm(source_type)
    if 'video' in value: return 'Video'
    if 'pdf' in value or 'document' in value or 'report' in value: return 'PDF'
    if 'social' in value or 'linkedin' in value or 'twitter' in value or 'facebook' in value or 'instagram' in value: return 'Social'
    if 'image' in value or 'photo' in value: return 'Image'
    return 'Article'

records = json.loads(TARGET.read_text(encoding='utf-8'))
by_url = {canonical(r.get('url')): r for r in records if canonical(r.get('url'))}
by_title_date = {(norm(r.get('title')), text(r.get('date'))): r for r in records if norm(r.get('title'))}

book = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
sheet = book['All Sources']
headers = [text(v) for v in next(sheet.iter_rows(values_only=True))]
added = matched = 0

for values in sheet.iter_rows(min_row=2, values_only=True):
    row = dict(zip(headers, values))
    title, url = text(row.get('Title/Headline')), text(row.get('URL'))
    if not title and not url:
        continue
    parsed_date = date_value(row.get('Date'))
    url_key = canonical(url)
    found = by_url.get(url_key) if url_key else None
    if not found:
        found = by_title_date.get((norm(title), parsed_date))
    if found:
        matched += 1
        found['duplicateCount'] = int(found.get('duplicateCount', 1)) + 1
        datasets = [x.strip() for x in text(found.get('sourceDataset')).split(';') if x.strip()]
        if LABEL not in datasets:
            found['sourceDataset'] = '; '.join(datasets + [LABEL])
        note = 'Matched ULTIMATE workbook by canonical URL' if url_key and canonical(found.get('url')) == url_key else 'Matched ULTIMATE workbook by normalized title and date'
        if note not in text(found.get('mergeNotes')):
            found['mergeNotes'] = '; '.join(filter(None, [text(found.get('mergeNotes')), note]))
        continue

    source_type = text(row.get('Source Type')) or 'Web source'
    year = year_value(row.get('Year'), parsed_date)
    supplied = text(row.get('Verification Status'))
    record = {
        'id': '', 'date': parsed_date, 'year': year,
        'type': broad_type(source_type), 'format': source_type,
        'publisher': text(row.get('Platform/Channel')) or 'Publisher not recorded',
        'title': title or 'Untitled source', 'language': text(row.get('Language')) or 'Unknown',
        'presence': text(row.get('Mention Type')) or 'MCCIA / Prashant Girbane occurrence reported',
        'topic': 'MCCIA / Prashant Girbane media monitoring',
        'description': text(row.get('Brief Description')),
        'status': 'Unverified' if 'unverified' in supplied.lower() else 'Partially verified',
        'url': url or None, 'mediaUrl': None,
        'notes': f'Imported from ULTIMATE Research workbook. Supplied verification label: {supplied or "not stated"}; independent review pending.',
        'sourceDataset': LABEL, 'duplicateCount': 1,
        'mergeNotes': 'New record absent from prior 424-record master'
    }
    records.append(record)
    if url_key: by_url[url_key] = record
    by_title_date[(norm(title), parsed_date)] = record
    added += 1

records.sort(key=lambda r: (text(r.get('date')), text(r.get('title'))), reverse=True)
for index, record in enumerate(records, 1):
    record['id'] = f'PG{index:03d}'

TARGET.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding='utf-8')
comparison = json.loads(COMPARE.read_text(encoding='utf-8'))
comparison['ultimate_workbook'] = {'source': LABEL, 'input_rows': added + matched, 'added': added, 'matched': matched}
comparison['prior_master_records'] = 424
comparison['final_records'] = len(records)
comparison['with_url'] = sum(bool(r.get('url')) for r in records)
comparison['without_url'] = sum(not r.get('url') for r in records)
comparison['verified'] = sum(r.get('status') == 'Verified' for r in records)
comparison['partial'] = sum(r.get('status') == 'Partially verified' for r in records)
comparison['unverified'] = sum(r.get('status') == 'Unverified' for r in records)
comparison['duplicates_merged'] = sum(max(0, int(r.get('duplicateCount', 1)) - 1) for r in records)
COMPARE.write_text(json.dumps(comparison, ensure_ascii=False, indent=2), encoding='utf-8')
print(json.dumps(comparison, ensure_ascii=False, indent=2))
