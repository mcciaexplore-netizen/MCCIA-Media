import json, re
from pathlib import Path
import openpyxl
from pypdf import PdfReader

files = [
    Path(r"C:\Users\Aarushi Gupta\Downloads\1787634550790-Prashant_Girbane_Comprehensive_Research_2018_2026.xlsx"),
    Path(r"C:\Users\Aarushi Gupta\Downloads\Mccia Media.xlsx"),
]
pdf_path = Path(r"C:\Users\Aarushi Gupta\Downloads\1787633727991-MCCIA_Comprehensive_Media_Report_2018_2026.pdf")
result = {"workbooks": [], "pdf": {}}

for path in files:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    info = {"path": str(path), "sheets": []}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        preview = [[str(v)[:300] if v is not None else None for v in row] for row in rows[:8]]
        urls = sorted(set(re.findall(r"https?://[^\s\)\]\>\"]+", "\n".join(str(v) for row in rows for v in row if v))))
        info["sheets"].append({"name": ws.title, "rows": len(rows), "cols": ws.max_column, "preview": preview, "urls": urls})
    result["workbooks"].append(info)

reader = PdfReader(str(pdf_path))
page_text = [page.extract_text() or "" for page in reader.pages]
full_text = "\n".join(page_text)
result["pdf"] = {
    "path": str(pdf_path), "pages": len(reader.pages), "chars": len(full_text),
    "urls": sorted(set(re.findall(r"https?://[^\s\)\]\>]+", full_text))),
    "page_previews": [text[:1200] for text in page_text],
}

out = Path(__file__).resolve().parents[1] / "source_inventory.json"
out.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps({"output": str(out), "workbooks": [{"name": Path(w['path']).name, "sheets": [(s['name'], s['rows'], len(s['urls'])) for s in w['sheets']]} for w in result['workbooks']], "pdf_pages": len(reader.pages), "pdf_urls": len(result['pdf']['urls'])}, ensure_ascii=False, indent=2))
