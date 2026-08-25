import json
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r'C:\Users\Aarushi Gupta\Downloads\News Paper data Files 2018-2026..2.xlsx')
TARGET = ROOT / 'app' / 'records.json'
COMPARE = ROOT / 'comparison_report.json'
LABEL = 'News Paper data Files 2018-2026 workbook'


def text(value):
    return str(value).strip() if value is not None else ''


def norm(value):
    return re.sub(r'[^a-z0-9\u0900-\u097f]+', ' ', text(value).lower()).strip()


def integer(value):
    match = re.search(r'\d+', text(value))
    return int(match.group()) if match else None


def parsed_date(day, month, year):
    day_number, year_number = integer(day), integer(year)
    if not day_number or not year_number:
        return ''
    month_text = text(month)
    for fmt in ('%B', '%b'):
        try:
            month_number = datetime.strptime(month_text, fmt).month
            return date(year_number, month_number, day_number).isoformat()
        except (ValueError, TypeError):
            pass
    month_number = integer(month)
    try:
        return date(year_number, month_number, day_number).isoformat() if month_number else ''
    except ValueError:
        return ''


records = json.loads(TARGET.read_text(encoding='utf-8'))
by_title_date = {(norm(r.get('title')), text(r.get('date'))): r for r in records if norm(r.get('title'))}

book = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
sheet = book['Sheet1']
headers = [text(v) for v in next(sheet.iter_rows(values_only=True))]
added = matched = skipped = 0

for values in sheet.iter_rows(min_row=2, values_only=True):
    row = dict(zip(headers, values))
    title = text(row.get('Title'))
    publisher = text(row.get('Newspaper'))
    year = integer(row.get('Year'))
    if not title or not year or year < 2018 or year > 2026:
        skipped += 1
        continue
    published = parsed_date(row.get('Date'), row.get('Month'), row.get('Year'))
    key = (norm(title), published)
    found = by_title_date.get(key)
    if found:
        matched += 1
        found['duplicateCount'] = int(found.get('duplicateCount', 1)) + 1
        datasets = [x.strip() for x in text(found.get('sourceDataset')).split(';') if x.strip()]
        if LABEL not in datasets:
            found['sourceDataset'] = '; '.join(datasets + [LABEL])
        note = 'Matched newspaper workbook by normalized title and publication date'
        if note not in text(found.get('mergeNotes')):
            found['mergeNotes'] = '; '.join(filter(None, [text(found.get('mergeNotes')), note]))
        continue

    supplement = text(row.get('Supplement'))
    page = text(row.get('Page no.'))
    person = text(row.get('person mentioned in news'))
    detail_parts = [f'Newspaper: {publisher or "not recorded"}']
    if supplement:
        detail_parts.append(f'Supplement: {supplement}')
    if page:
        detail_parts.append(f'Page: {page}')
    record = {
        'id': '',
        'date': published,
        'year': year,
        'type': 'Article',
        'format': 'Print newspaper index',
        'publisher': publisher or 'Newspaper not recorded',
        'title': title,
        'language': text(row.get('Language')) or 'Unknown',
        'presence': person or 'MCCIA-related newspaper item; named person not recorded',
        'topic': 'MCCIA newspaper coverage',
        'description': '; '.join(detail_parts),
        'status': 'Unverified',
        'url': None,
        'mediaUrl': None,
        'notes': 'Imported from a supplied print-newspaper index. No public source URL or clipping was supplied; publication details require independent verification.',
        'sourceDataset': LABEL,
        'duplicateCount': 1,
        'mergeNotes': 'New print-index record absent from the prior dashboard dataset',
    }
    records.append(record)
    by_title_date[key] = record
    added += 1

records.sort(key=lambda r: (text(r.get('date')), text(r.get('title'))), reverse=True)
for index, record in enumerate(records, 1):
    record['id'] = f'PG{index:04d}'

TARGET.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding='utf-8')
comparison = json.loads(COMPARE.read_text(encoding='utf-8'))
comparison['newspaper_workbook'] = {
    'source': LABEL,
    'input_rows': sheet.max_row - 1,
    'added': added,
    'matched': matched,
    'skipped': skipped,
}
comparison['final_records'] = len(records)
comparison['with_url'] = sum(bool(r.get('url')) for r in records)
comparison['without_url'] = sum(not r.get('url') for r in records)
comparison['verified'] = sum(r.get('status') == 'Verified' for r in records)
comparison['partial'] = sum(r.get('status') == 'Partially verified' for r in records)
comparison['unverified'] = sum(r.get('status') == 'Unverified' for r in records)
comparison['duplicates_merged'] = sum(max(0, int(r.get('duplicateCount', 1)) - 1) for r in records)
COMPARE.write_text(json.dumps(comparison, ensure_ascii=False, indent=2), encoding='utf-8')
print(json.dumps(comparison['newspaper_workbook'] | {'final_records': len(records)}, ensure_ascii=False, indent=2))
